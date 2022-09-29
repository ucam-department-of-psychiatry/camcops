#!/usr/bin/env python

"""
camcops_server/cc_modules/tests/cc_spreadsheet_tests.py

===============================================================================

    Copyright (C) 2012, University of Cambridge, Department of Psychiatry.
    Created by Rudolf Cardinal (rnc1001@cam.ac.uk).

    This file is part of CamCOPS.

    CamCOPS is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    CamCOPS is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with CamCOPS. If not, see <https://www.gnu.org/licenses/>.

===============================================================================

"""

import io
from typing import Any, Dict
from unittest import TestCase
import uuid
from xml.dom.minidom import parseString
import zipfile

from camcops_server.cc_modules.cc_spreadsheet import (
    SpreadsheetCollection,
    SpreadsheetPage,
    XLSX_VIA_PYEXCEL,
)

if XLSX_VIA_PYEXCEL:
    import pyexcel_xlsx  # e.g. pip install pyexcel-xlsx==0.5.7

    openpyxl = XLWorkbook = XLWorksheet = None
else:
    import openpyxl
    from openpyxl.workbook.workbook import Workbook as XLWorkbook

    pyexcel_xlsx = None


# =============================================================================
# Unit tests
# =============================================================================


class SpreadsheetCollectionTests(TestCase):
    def test_xlsx_created_from_zero_rows(self) -> None:
        page = SpreadsheetPage(name="test", rows=[])
        coll = SpreadsheetCollection()
        coll.add_page(page)

        output = coll.as_xlsx()

        # https://en.wikipedia.org/wiki/List_of_file_signatures
        self.assertEqual(output[0], 0x50)
        self.assertEqual(output[1], 0x4B)
        self.assertEqual(output[2], 0x03)
        self.assertEqual(output[3], 0x04)

    def test_xlsx_worksheet_names_are_page_names(self) -> None:
        page1 = SpreadsheetPage(name="name 1", rows=[{"test data 1": "row 1"}])
        page2 = SpreadsheetPage(name="name 2", rows=[{"test data 2": "row 1"}])
        page3 = SpreadsheetPage(name="name 3", rows=[{"test data 3": "row 1"}])
        coll = SpreadsheetCollection()

        coll.add_pages([page1, page2, page3])

        data = coll.as_xlsx()
        buffer = io.BytesIO(data)
        expected_sheetnames = ["name 1", "name 2", "name 3"]
        if openpyxl:
            wb = openpyxl.load_workbook(buffer)  # type: XLWorkbook
            self.assertEqual(wb.sheetnames, expected_sheetnames)
        else:
            wb = pyexcel_xlsx.get_data(buffer)  # type: Dict[str, Any]
            sheetnames = list(wb.keys())
            self.assertEqual(sheetnames, expected_sheetnames)

    def test_xlsx_page_name_exactly_31_chars_not_truncated(self) -> None:
        page = SpreadsheetPage(
            name="abcdefghijklmnopqrstuvwxyz78901",
            rows=[{"test data 1": "row 1"}],
        )
        coll = SpreadsheetCollection()

        self.assertEqual(
            coll.get_sheet_title(page), "abcdefghijklmnopqrstuvwxyz78901"
        )

    def test_xlsx_page_name_over_31_chars_truncated(self) -> None:
        page = SpreadsheetPage(
            name="abcdefghijklmnopqrstuvwxyz78901234",
            rows=[{"test data 1": "row 1"}],
        )
        coll = SpreadsheetCollection()

        self.assertEqual(
            coll.get_sheet_title(page), "abcdefghijklmnopqrstuvwxyz78..."
        )

    def test_xlsx_invalid_chars_in_page_name_replaced(self) -> None:
        page = SpreadsheetPage(
            name="[a]b\\c:d/e*f?g'h", rows=[{"test data 1": "row 1"}]
        )
        coll = SpreadsheetCollection()

        self.assertEqual(coll.get_sheet_title(page), "_a_b_c_d_e_f_g_h")

    def test_ods_page_name_sanitised(self) -> None:
        # noinspection PyUnresolvedReferences
        page = SpreadsheetPage(
            name="What perinatal service have you accessed?",
            rows=[{"test data 1": "row 1"}],
        )
        coll = SpreadsheetCollection()
        coll.add_pages([page])

        data = coll.as_ods()

        zf = zipfile.ZipFile(io.BytesIO(data), "r")
        content = zf.read("content.xml")
        doc = parseString(content)
        sheets = doc.getElementsByTagName("table:table")
        self.assertEqual(
            sheets[0].getAttribute("table:name"),
            "What perinatal service have ...",
        )

    def test_worksheet_names_are_not_duplicated(self) -> None:
        page1 = SpreadsheetPage(
            name="abcdefghijklmnopqrstuvwxyz78901234",
            rows=[{"test data 1": "row 1"}],
        )
        page2 = SpreadsheetPage(
            name="ABCDEFGHIJKLMNOPQRSTUVWXYZ789012345",
            rows=[{"test data 2": "row 1"}],
        )
        page3 = SpreadsheetPage(
            name="abcdefghijklmnopqrstuvwxyz7890123456",
            rows=[{"test data 3": "row 1"}],
        )
        coll = SpreadsheetCollection()

        coll.add_pages([page1, page2, page3])

        valid_sheet_names = coll.get_pages_with_valid_sheet_names()

        names = [v for k, v in valid_sheet_names.items()]

        self.assertIn("abcdefghijklmnopqrstuvwxyz78...", names)
        self.assertIn("ABCDEFGHIJKLMNOPQRSTUVWXYZ78..1", names)
        self.assertIn("abcdefghijklmnopqrstuvwxyz78..2", names)

    def test_uuid_exported_to_ods_as_string(self) -> None:
        test_uuid = uuid.UUID("6457cb90-1ca0-47a7-9f40-767567819bee")

        page = SpreadsheetPage(name="Testing", rows=[{"UUID": test_uuid}])
        coll = SpreadsheetCollection()
        coll.add_pages([page])

        data = coll.as_ods()
        zf = zipfile.ZipFile(io.BytesIO(data), "r")
        content = zf.read("content.xml")
        doc = parseString(content)
        text_values = [
            t.firstChild.nodeValue for t in doc.getElementsByTagName("text:p")
        ]

        self.assertIn("UUID", text_values)
        self.assertIn("6457cb90-1ca0-47a7-9f40-767567819bee", text_values)

    def test_uuid_exported_to_xlsx_as_string(self) -> None:
        test_uuid = uuid.UUID("6457cb90-1ca0-47a7-9f40-767567819bee")

        page = SpreadsheetPage(name="Testing", rows=[{"UUID": test_uuid}])
        coll = SpreadsheetCollection()
        coll.add_pages([page])

        data = coll.as_xlsx()
        buffer = io.BytesIO(data)
        if openpyxl:
            self.fail("This test has not been written for openpyxl")
        else:
            wb = pyexcel_xlsx.get_data(buffer)  # type: Dict[str, Any]
            self.assertIn(["UUID"], wb["Testing"])
            self.assertIn(
                ["6457cb90-1ca0-47a7-9f40-767567819bee"], wb["Testing"]
            )
