#!/usr/bin/env python

"""
camcops_server/cc_modules/cc_tsv.py

===============================================================================

    Copyright (C) 2012-2019 Rudolf Cardinal (rudolf@pobox.com).

    This file is part of CamCOPS.

    CamCOPS is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    CamCOPS is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with CamCOPS. If not, see <http://www.gnu.org/licenses/>.

===============================================================================

**Helper functions/classes for spreadsheet-style tab-separated value (TSV)
exports.**

"""

from collections import OrderedDict
import csv
import io
import logging
import re
from typing import Any, Dict, Iterable, List, Optional, Union
from unittest import TestCase
import zipfile

from cardinal_pythonlib.excel import (
    convert_for_openpyxl,
    excel_to_bytes,
)
from cardinal_pythonlib.logs import BraceStyleAdapter
from odswriter import ODSWriter, Sheet as ODSSheet
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook as XLWorkbook
from openpyxl.worksheet.worksheet import Worksheet as XLWorksheet
# from pyexcel_ods3 import save_data  # poor; use odswriter

log = BraceStyleAdapter(logging.getLogger(__name__))


# =============================================================================
# TSV output holding structures
# =============================================================================

class TsvPage(object):
    """
    Represents a single TSV "spreadsheet".
    """
    def __init__(self, name: str,
                 rows: List[Union[Dict[str, Any], OrderedDict]]) -> None:
        """
        Args:
            name: name for the whole sheet
            rows: list of rows, where each row is a dictionary mapping
                column name to value
        """
        assert name, "Missing name"
        self.name = name
        self.rows = rows
        self.headings = []  # type: List[str]
        for row in rows:
            self._add_headings_if_absent(row.keys())

    def __str__(self) -> str:
        return f"TsvPage: name={self.name}\n{self.get_tsv()}"

    @property
    def empty(self) -> bool:
        """
        Do we have zero rows?
        """
        return len(self.rows) == 0

    def _add_headings_if_absent(self, headings: Iterable[str]) -> None:
        """
        Add any headings we've not yet seen to our list of headings.
        """
        for h in headings:
            if h not in self.headings:
                self.headings.append(h)

    def add_or_set_value(self, heading: str, value: Any) -> None:
        """
        If we contain only a single row, this function will set the value
        for a given column (``heading``) to ``value``.

        Raises:
            :exc:`AssertionError` if we don't have exactly 1 row
        """
        assert len(self.rows) == 1, "add_value can only be used if #rows == 1"
        self._add_headings_if_absent([heading])
        self.rows[0][heading] = value

    def add_or_set_column(self, heading: str, values: List[Any]) -> None:
        """
        Set the column labelled ``heading`` so it contains the values specified
        in ``values``. The length of ``values`` must equal the number of rows
        that we already contain.

        Raises:
            :exc:`AssertionError` if the number of values doesn't match
            the number of existing rows
        """
        assert len(values) == len(self.rows), "#values != #existing rows"
        self._add_headings_if_absent([heading])
        for i, row in enumerate(self.rows):
            row[heading] = values[i]

    def add_or_set_columns_from_page(self, other: "TsvPage") -> None:
        """
        This function presupposes that ``self`` and ``other`` are two pages
        ("spreadsheets") with *matching* rows.

        It updates values or creates columns in ``self`` such that the values
        from all columns in ``other`` are written to the corresponding rows of
        ``self``.

        Raises:
            :exc:`AssertionError` if the two pages (sheets) don't have
            the same number of rows.
        """
        assert len(self.rows) == len(other.rows), "Mismatched #rows"
        self._add_headings_if_absent(other.headings)
        for i, row in enumerate(self.rows):
            for k, v in other.rows[i].items():
                row[k] = v

    def add_rows_from_page(self, other: "TsvPage") -> None:
        """
        Add all rows from ``other`` to ``self``.
        """
        self._add_headings_if_absent(other.headings)
        self.rows.extend(other.rows)

    def sort_headings(self) -> None:
        """
        Sort our headings internally.
        """
        self.headings.sort()

    def get_tsv(self, dialect: str = "excel-tab") -> str:
        r"""
        Returns the entire page (sheet) as TSV: one header row and then
        lots of data rows.

        For the dialect, see
        https://docs.python.org/3/library/csv.html#csv.excel_tab.

        For CSV files, see RGC 4180: https://tools.ietf.org/html/rfc4180.

        For TSV files, see
        https://www.iana.org/assignments/media-types/text/tab-separated-values.

        Test code:

        .. code-block:: python

            import io
            import csv
            from typing import List

            def test(row: List[str], dialect: str = "excel-tab") -> str:
                f = io.StringIO()
                writer = csv.writer(f, dialect=dialect)
                writer.writerow(row)
                return f.getvalue()

            test(["hello", "world"])
            test(["hello\ttab", "world"])  # actual tab within double quotes
            test(["hello\nnewline", "world"])  # actual newline within double quotes
            test(['hello"doublequote', "world"])  # doubled double quote within double quotes

        """  # noqa
        f = io.StringIO()
        writer = csv.writer(f, dialect=dialect)
        writer.writerow(self.headings)
        for row in self.rows:
            writer.writerow([row.get(h) for h in self.headings])
        return f.getvalue()

    def write_to_xlsx_worksheet(self, ws: XLWorksheet) -> None:
        """
        Writes data from this page to an existing XLSX worksheet.
        """
        ws.append(self.headings)
        for row in self.rows:
            ws.append([convert_for_openpyxl(row.get(h))
                       for h in self.headings])

    def write_to_ods_worksheet(self, ws: ODSSheet) -> None:
        """
        Writes data from this page to an existing ODS sheet.
        """
        ws.writerow(self.headings)
        for row in self.rows:
            ws.writerow([row.get(h) for h in self.headings])


class TsvCollection(object):
    """
    A collection of :class:`camcops_server.cc_modules.cc_tsv.TsvPage` pages
    (spreadsheets), like an Excel workbook.
    """
    def __init__(self) -> None:
        self.pages = []  # type: List[TsvPage]

    def __str__(self) -> str:
        return (
            "TsvCollection:\n" +
            "\n\n".join(page.get_tsv() for page in self.pages)
        )

    def page_with_name(self, page_name: str) -> Optional[TsvPage]:
        """
        Returns the page with the specific name, or ``None`` if no such
        page exists.
        """
        return next((page for page in self.pages if page.name == page_name),
                    None)

    def add_page(self, page: TsvPage) -> None:
        """
        Adds a new page to our collection. If the new page has the same name
        as an existing page, rows from the new page are added to the existing
        page. Does nothing if the new page is empty.
        """
        if page.empty:
            return
        existing_page = self.page_with_name(page.name)
        if existing_page:
            # Blend with existing page
            existing_page.add_rows_from_page(page)
        else:
            # New page
            self.pages.append(page)

    def add_pages(self, pages: List[TsvPage]) -> None:
        """
        Adds all ``pages`` to our collection, via :func:`add_page`.
        """
        for page in pages:
            self.add_page(page)

    def sort_headings_within_all_pages(self) -> None:
        """
        Sort headings within each of our pages.
        """
        for page in self.pages:
            page.sort_headings()

    def sort_pages(self) -> None:
        """
        Sort our pages by their page name.
        """
        self.pages.sort(key=lambda p: p.name)

    def get_page_names(self) -> List[str]:
        """
        Return a list of the names of all our pages.
        """
        return [p.name for p in self.pages]

    def get_tsv_file(self, page_name: str) -> str:
        """
        Returns a TSV file for a named page.

        Raises:
            :exc:`AssertionError` if the named page does not exist

        """
        page = self.page_with_name(page_name)
        assert page is not None, f"No such page with name {page_name}"
        return page.get_tsv()

    def as_zip(self, encoding: str = "utf-8") -> bytes:
        """
        Returns the TSV collection as a ZIP file containing TSV files.

        Args:
            encoding: encoding to use when writing the TSV files
        """
        with io.BytesIO() as memfile:
            with zipfile.ZipFile(memfile, "w") as z:
                # Write to ZIP.
                # If there are no valid task instances, there'll be no TSV;
                # that's OK.
                for filename_stem in self.get_page_names():
                    tsv_filename = filename_stem + ".tsv"
                    tsv_contents = self.get_tsv_file(page_name=filename_stem)
                    z.writestr(tsv_filename, tsv_contents.encode(encoding))
            zip_contents = memfile.getvalue()
        return zip_contents

    def as_xlsx(self) -> bytes:
        """
        Returns the TSV collection as an XLSX (Excel) file.
        """
        # Marginal performance gain with write_only. Does not automatically
        # add a blank sheet
        wb = XLWorkbook(write_only=True)

        valid_name_dict = self.get_pages_with_valid_sheet_names()
        for page, title in valid_name_dict.items():
            ws = wb.create_sheet(title=title)
            page.write_to_xlsx_worksheet(ws)

        return excel_to_bytes(wb)

    @staticmethod
    def get_sheet_title(page: TsvPage) -> str:
        # See openpyxl/workbook/child.py

        # Excel prohibits \,*,?,:,/,[,]
        # LibreOffice also prohibits ' as first or last character but let's
        # just replace that globally
        title = re.sub(r"[\\*?:/\[\]']", "_", page.name)

        if len(title) > 31:
            title = f"{title[:28]}..."

        return title

    def as_ods(self) -> bytes:
        """
        Returns the TSV collection as an ODS (OpenOffice spreadsheet document)
        file.
        """
        with io.BytesIO() as memfile:
            with ODSWriter(memfile) as odsfile:
                valid_name_dict = self.get_pages_with_valid_sheet_names()
                for page, title in valid_name_dict.items():
                    sheet = odsfile.new_sheet(name=title)
                    page.write_to_ods_worksheet(sheet)
            contents = memfile.getvalue()
        return contents

    def get_pages_with_valid_sheet_names(self) -> Dict[TsvPage, str]:
        name_dict = {}

        for page in self.pages:
            name_dict[page] = self.get_sheet_title(page)

        self.make_sheet_names_unique(name_dict)

        return name_dict

    @staticmethod
    def make_sheet_names_unique(name_dict: Dict[TsvPage, str]) -> None:
        # See also avoid_duplicate_name in openpxl/workbook/child.py
        # We keep the 31 character restriction

        unique_names = []

        for page, name in name_dict.items():
            attempt = 0

            while name.lower() in unique_names:
                attempt += 1

                if attempt > 1000:
                    # algorithm failure, better to let Excel deal with the
                    # consequences than get stuck in a loop
                    log.debug(
                        f"Failed to generate a unique sheet name from {name}"
                    )
                    break

                match = re.search(r'\d+$', name)
                count = 0
                if match is not None:
                    count = int(match.group())

                new_suffix = str(count + 1)
                name = name[:-len(new_suffix)] + new_suffix
            name_dict[page] = name
            unique_names.append(name.lower())


# =============================================================================
# Unit tests
# =============================================================================

class TsvCollectionTests(TestCase):
    def test_xlsx_created_from_zero_rows(self) -> None:
        page = TsvPage(name="test", rows=[])
        coll = TsvCollection()
        coll.add_page(page)

        output = coll.as_xlsx()

        # https://en.wikipedia.org/wiki/List_of_file_signatures
        self.assertEqual(output[0], 0x50)
        self.assertEqual(output[1], 0x4B)
        self.assertEqual(output[2], 0x03)
        self.assertEqual(output[3], 0x04)

    def test_xlsx_worksheet_names_are_page_names(self) -> None:
        page1 = TsvPage(name="name 1",
                        rows=[{"test data 1": "row 1"}])
        page2 = TsvPage(name="name 2",
                        rows=[{"test data 2": "row 1"}])
        page3 = TsvPage(name="name 3",
                        rows=[{"test data 3": "row 1"}])
        coll = TsvCollection()

        coll.add_pages([page1, page2, page3])

        data = coll.as_xlsx()
        buffer = io.BytesIO(data)
        wb = load_workbook(buffer)
        self.assertEqual(
            wb.sheetnames,
            [
                "name 1",
                "name 2",
                "name 3",
            ]
        )

    def test_xlsx_page_name_exactly_31_chars_not_truncated(self) -> None:
        page = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "abcdefghijklmnopqrstuvwxyz78901"
        )

    def test_xlsx_page_name_over_31_chars_truncated(self) -> None:
        page = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901234",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "abcdefghijklmnopqrstuvwxyz78..."
        )

    def test_xlsx_invalid_chars_in_page_name_replaced(self) -> None:
        page = TsvPage(name="[a]b\\c:d/e*f?g'h",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()

        self.assertEqual(
            coll.get_sheet_title(page),
            "_a_b_c_d_e_f_g_h"
        )

    def test_ods_page_name_sanitised(self) -> None:
        import xml.dom.minidom
        page = TsvPage(name="What perinatal service have you accessed?",
                       rows=[{"test data 1": "row 1"}])
        coll = TsvCollection()
        coll.add_pages([page])

        data = coll.as_ods()

        zf = zipfile.ZipFile(io.BytesIO(data), "r")
        content = zf.read('content.xml')
        doc = xml.dom.minidom.parseString(content)
        sheets = doc.getElementsByTagName('table:table')
        self.assertEqual(sheets[0].getAttribute("table:name"),
                         "What perinatal service have ...")

    def test_worksheet_names_are_not_duplicated(self) -> None:
        page1 = TsvPage(name="abcdefghijklmnopqrstuvwxyz78901234",
                        rows=[{"test data 1": "row 1"}])
        page2 = TsvPage(name="ABCDEFGHIJKLMNOPQRSTUVWXYZ789012345",
                        rows=[{"test data 2": "row 1"}])
        page3 = TsvPage(name="abcdefghijklmnopqrstuvwxyz7890123456",
                        rows=[{"test data 3": "row 1"}])
        coll = TsvCollection()

        coll.add_pages([page1, page2, page3])

        valid_sheet_names = coll.get_pages_with_valid_sheet_names()

        names = [v for k, v in valid_sheet_names.items()]

        self.assertIn("abcdefghijklmnopqrstuvwxyz78...", names)
        self.assertIn("ABCDEFGHIJKLMNOPQRSTUVWXYZ78..1", names)
        self.assertIn("abcdefghijklmnopqrstuvwxyz78..2", names)
